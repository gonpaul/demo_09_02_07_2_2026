from django.core.management.base import BaseCommand
import openpyxl
from shop.models import CustomUser, Postavschik, Proizvoditel, KategoriaTovara, PunktVidachi, StatusZakaza, Tovar, TovarVZakaze, Zakaz

class Command(BaseCommand):
    help = 'импорт данных из экселя'

    def handle(self, *args, **kwargs):
        wb = openpyxl.load_workbook('/home/user/Documents/hexlet/exam/shoes_var/imports/user_import.xlsx')
        wb = wb.active
        for row in wb.iter_rows(min_row=2, values_only=True):
            print(row)
            if row[0]:
                role, fio, email, raw_password = row
                user, created = CustomUser.objects.get_or_create(
                    username = email,
                    role = role,
                    fio = fio
                )
                user.set_password(raw_password)
                user.save()
                print('Сохранён пользователь: ', user)
        
        # wb = openpyxl.load_workbook('/home/user/Documents/hexlet/exam/shoes_var/imports/Tovar.xlsx')
        # wb = wb.active
        # for row in wb.iter_rows(min_row=2, values_only=True):
        #     print(row)
        #     if row[0]:
        #         artikul, name, unit, cena, postavschik, proizvoditel, kategoria, skidka, na_sklade, opisanie, foto = row
        #         postavschik, created = Postavschik.objects.get_or_create(name=postavschik) 
        #         proizvoditel, created = Proizvoditel.objects.get_or_create(name=proizvoditel)
        #         kategoria, created = KategoriaTovara.objects.get_or_create(name=kategoria)
        #         # print(postavschik)

        #         tovar, created = Tovar.objects.get_or_create(
        #             artikul = artikul,
        #             name = name,
        #             cena = cena,
        #             postavschik = postavschik,
        #             proizvoditel = proizvoditel,
        #             kategoria = kategoria,
        #             skidka = skidka,
        #             na_sklade = na_sklade,
        #             opisanie = opisanie,
        #             foto = foto
        #         )
        #         print('Сохранённый товар: ', tovar)

        # wb = openpyxl.load_workbook('/home/user/Documents/hexlet/exam/shoes_var/imports/Пункты выдачи_import.xlsx')
        # wb = wb.active
        # for row in wb.iter_rows(min_row=2, values_only=True):
        #     if row[0]:
        #         print(row)
        #         pv, created = PunktVidachi.objects.get_or_create(name=row[0])
        #         print('Created punkt: ', pv)

        # wb = openpyxl.load_workbook('/home/user/Documents/hexlet/exam/shoes_var/imports/Заказ_import.xlsx')
        # wb = wb.active
        # for row in wb.iter_rows(min_row=2, values_only=True):
        #     if row[0]:
        #         print(row)
        #         id, artikuls, data_zakaza, data_dostavki, punkt_vidachi, fio_klienta, kod, status = row
        #         status, created = StatusZakaza.objects.get_or_create(name=status)
        #         zakaz, created = Zakaz.objects.get_or_create(
        #             id = id,
        #             data_zakaza = data_zakaza,
        #             data_dostavki = data_dostavki,
        #             punkt_vidachi = PunktVidachi.objects.get(id=punkt_vidachi),
        #             fio_klienta = fio_klienta,
        #             kod = kod,
        #             status = status,
        #         )
        #         print('Создан заказ', zakaz)

        #         splitted = list(map(lambda x: x.strip(), artikuls.split(',')))
        #         print(splitted)
        #         artikul_one, kolvo_one = splitted[0], splitted[1]
        #         artikul_two, kolvo_two = splitted[2], splitted[3]

        #         tovar_vzakaze_one = TovarVZakaze.objects.get_or_create(
        #             zakaz=zakaz,
        #             tovar=Tovar.objects.get(artikul=artikul_one),
        #             kolichestvo=kolvo_one
        #         )
        #         tovar_vzakaze_two = TovarVZakaze.objects.get_or_create(
        #             zakaz=zakaz,
        #             tovar=Tovar.objects.get(artikul=artikul_two),
        #             kolichestvo=kolvo_two
        #         )
        #         print('Созданы единицы заказа ', tovar_vzakaze_one, tovar_vzakaze_two)





